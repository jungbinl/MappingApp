import os
import glob
import base64
import openpyxl
from openai import OpenAI

# OpenAI 클라이언트 초기화
client = OpenAI(api_key="")

def encode_image_to_base64(image_path):
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

def extract_single_image_data(image_path, image_label):
    base64_image = encode_image_to_base64(image_path)
    print(f"🧠 [{image_label}] GPT-4o가 사진 분석 및 데이터 추출 중...")
    
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "너는 현장 수기 서류를 정밀하게 판독하여 매칭하는 데이터 전문가야.\n"
                            "이 사진 속에는 인쇄된 영문+숫자 혼합 코드(예: SF1, SD-3, sd1, SD1, NF1, SDX1 등)와\n"
                            "그에 매칭되는 손글씨 숫자들이 무작위 양식으로 적혀 있어.\n\n"
                            "다음 규칙을 엄격히 적용하여 완벽한 JSON 데이터 객체를 만들어줘:\n"
                            "1. [코드 자동 인식]: 사진에 등장하는 모든 영문+숫자 조합의 식별 코드를 눈에 보이는 대로 전부 수집하여 Key로 잡아줘.\n"
                            "2. 소문자로 적힌 코드(예: sd-3, sd1)가 있다면 대문자 형태(SD-3, SD1)로 변환하거나 보이는 형태를 유지하여 빠짐없이 추출할 것.\n"
                            "3. 가려졌거나 우측에 손글씨 숫자가 없는 빈칸 코드는 값(Value)을 빈 문자열(\"\")로 처리할 것.\n"
                            "4. 취소선(낙서)이 쳐진 숫자는 제외하고, 최종 수정 기입된 유효한 숫자 데이터만 깔끔하게 추출할 것.\n"
                            "5. 슬래시(/)나 하이픈(-)이 포함된 수기 일련번호나 복수 데이터도 보이는 그대로 문자열로 저장할 것.\n\n"
                            "6. SD의 경우는 SD-1, SD-2, SD-3 이 있으니 주의하도록"
                            "반드시 다른 헛소리는 제외하고 아래의 순수한 JSON 구조로만 답변해야 해:\n"
                            "{\n"
                            "  \"인식된코드1\": \"수기숫자\",\n"
                            "  \"인식된코드2\": \"\"\n"
                            "}"
                        )
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    }
                ]
            }
        ],
        max_tokens=1500,
        temperature=0.0
    )
    return response.choices[0].message.content

def save_multiple_to_excel(results_dict, output_path):
    wb = openpyxl.load_workbook(output_path)

    first_sheet = True
    
    for image_label, gpt_text in results_dict.items():
        upper_text = gpt_text.upper()
        
        if "SF" in upper_text:
            current_sheet_title = "SF"
        elif "SD" in upper_text:
            current_sheet_title = "SD"
        elif "NF" in upper_text:
            current_sheet_title = "NF"

        if current_sheet_title in wb.sheetnames:
            print(f"♻️  [{current_sheet_title}] 시트가 이미 존재하여 내부 내용을 지우고 덮어씌웁니다.")
            ws = wb[current_sheet_title]
            
            ws.delete_rows(1, 200)
        else:
            # 새로 생성한 파일이고 첫 루프라면, 자동 생성된 임시 빈 시트를 재활용합니다.
            if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "__TEMP_EMPTY_SHEET__":
                ws = wb.active
                ws.title = current_sheet_title
            else:
                ws = wb.create_sheet(title=current_sheet_title)
            
        ws.append(["코드 (Code)", "작성된 숫자 1 (Value 1)", "작성된 숫자 2 (Value 2)"])
        
        lines = gpt_text.strip().split('\n')

        for line in lines:
            if ":" in line:
                parts = line.split(":", 1)
                
                code = parts[0].replace("*", "").strip()
                value = parts[1].strip()

                code = code.strip(' "\'')
                value = value.strip(' "\'')

                if code.endswith(','): code = code[:-1].strip()
                if value.endswith(','): value = value[:-1].strip()

                code = code.strip(' "\'')
                value = value.strip(' "\'')

                # 분리될 값을 저장할 변수 초기화
                val1 = None
                val2 = None

                if value != "":
                    # 💡 [핵심 추가] 숫자와 숫자 사이에 슬래시(/)가 있는지 검사합니다.
                    if '/' in value:
                        sub_parts = value.split('/')
                        raw_val1 = sub_parts[0].strip()
                        raw_val2 = sub_parts[1].strip()
                    else:
                        raw_val1 = value
                        raw_val2 = ""

                    if raw_val1 != "":
                        try:
                            val1 = int(raw_val1)
                        except ValueError:
                            val1 = raw_val1

                    if raw_val2 != "":
                        try:
                            val2 = int(raw_val2)
                        except ValueError:
                            val2 = raw_val2 

                ws.append([code, val1, val2])
    wb.save(output_path)
    print("-" * 50)
    print(f"✨ [통합 변환 완료] 폴더 내 모든 사진 데이터가 하나의 엑셀 파일로 묶였습니다!")
    print(f"📄 저장 경로: {os.path.abspath(output_path)}")
    print("-" * 50)

# --- [실행부] ---
if __name__ == "__main__":
    input_folder = "C:/Users/JUNGBIN LEE/Desktop/scan"
    final_output = "C:/Users/JUNGBIN LEE/Desktop/main.xlsx"
    

        # 폴더 내의 jpg, jpeg, png 파일들을 전부 긁어옵니다.
    image_extensions = ["*.jpg", "*.jpeg", "*.png"]
    image_list = []
    for ext in image_extensions:
        image_list.extend(glob.glob(os.path.join(input_folder, ext)))
            
    print(f"📸 총 {len(image_list)}개의 이미지를 발견했습니다. 분석을 시작합니다.")
    all_results = {}
            
    # 발견된 이미지들을 순서대로 루프 작동
    for idx, img_path in enumerate(image_list, start=1):
        # 파일명을 탭 이름으로 쓰거나 '사진_1' 형태로 지정
        filename = os.path.basename(img_path)
        label = f"사진_{idx}_{filename[:10]}" # 파일명 앞 10자리를 탭이름으로 활용
                
        raw_text = extract_single_image_data(img_path, label)
        all_results[label] = raw_text
                
            # 엑셀로 병합 저장
    save_multiple_to_excel(all_results, final_output)
    save_multiple_to_excel(all_results, "C:/Users/JUNGBIN LEE/Desktop/main.xlsx")